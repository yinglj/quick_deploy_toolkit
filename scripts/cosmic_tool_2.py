#!/usr/bin/env python
# -*- coding: utf-8 -*-

'''
name:
cosmic_tool_1.py
email:
yinglj@asiainfo.com
title:
cosmic_tool_1.py for Linux
'''

import os
import sys
import openpyxl
from progressbar import *

class CCosmic:
    def __init__(self):
        self.listProcessFiles = []
        self.wb = None
        self.ws = None
        #self.wb = openpyxl.Workbook()
        #self.ws = self.wb.create_sheet('Sheet1', index=0)
        self.catalog_names = {} # 3.2.10.1.1. MDB功能类型录入 => key: 3.2.10.1.1, value: MDB功能类型录入
        self.catalog_sections = {} # 【业务定义】=> key:MDB功能类型录入, value:{key:业务定义, value:[]}
        pass
                
    def processAllFiles(self):
        if not os.path.exists(sys.argv[1]):
            print('Excel template file(%s) not exists.' %(sys.argv[1]))
            return False
        self.wb = openpyxl.load_workbook(sys.argv[1])
        self.ws = self.wb.active
        for i in range(2, len(sys.argv)):
            self.listProcessFiles.append(sys.argv[i])
        for v in self.listProcessFiles:
            self.processOneFile(v)
        return True
            
    def processOneFile(self, txt_file):
        if not os.path.exists(txt_file):
            print('%s file not exist.' %(txt_file))
            return False
        line_no = 0
        catalog_begin = False
        section_begin = False
        catalog_name = ''
        section_name = ''
        self.catalog_names.clear();
        self.catalog_sections.clear();
        for line in open(txt_file, encoding='UTF-8'):
            line_no = line_no + 1
            line = line.strip()
            if len(line) == 0:
                continue
            if line[0].isdigit():
                catalog_name = line
                line = line.split('.')
                self.catalog_names['.'.join(line[0:-1])] = line[-1]
                self.catalog_sections[catalog_name] = {}
                catalog_begin = True
                section_begin = False
                continue
            if catalog_begin:
                if line[0] in ('[', '【') and line[-1] in (']', '】'):
                    section_begin = True
                    section_name = line[1:-1]
                    self.catalog_sections[catalog_name][section_name] = []
                elif section_begin:
                    self.catalog_sections[catalog_name][section_name].append(line)
                else:
                    print('file:%s, line:%d, "%s" format error.' %(txt_file, line_no, line))
                    return False
        if not catalog_begin:
            print('%s file format error.' %(txt_file))
            return False  
        #self.writeExcelHead()
        self.writeExcel()                
        return True

    #复制单元格格式
    def duplicateCellStyle(self, cell1, cell2):
        from copy import copy
        cell1.font = copy(cell2.font)
        cell1.border = copy(cell2.border)
        cell1.fill = copy(cell2.fill)
        cell1.number_format = copy(cell2.number_format)
        cell1.protection = copy(cell2.protection)
        cell1.alignment = copy(cell2.alignment)
    
    #复制整行单元格格式，从row行复制到row+1行
    def duplicateRowStyle(self, ws, row, seqno=1):
        for i in range(ws.min_column-1, ws.max_column-1):
            cell1 = '{}{}'.format(chr(ord('A')+i), row[0].row+seqno)
            self.duplicateCellStyle(ws[cell1], row[i])   #复制单元格格式
                
    def writeExcelHead(self):
        tile_names = ['客户需求','功能用户','功能用户需求','触发事件','功能过程','子过程描述','数据移动类型','数据组','数据属性','CFP','ΣCFP']
        for idx, tile in enumerate(tile_names):
            cell = '{}{}'.format(chr(ord('A')+idx), 1)
            self.ws[cell] = tile
         
    def writeExcel(self):
        sect_names = ['业务定义', '功能要求', '业务要素']
        #write_row_cnt = 1
        write_row_cnt = self.ws.max_row
        in_elem = ''
        out_elem = ''
        data_group = ''
        data_type = ['E','R','R','X','W']
        func_desc = ['传入{}结构参数', '读取并解析传入{}结构参数', '{}逻辑进行封装处理', '{}结果封装并进行校验','将{}记录到日志文件中']
        data_group_desc = ['{}输入信息','{}校验信息','{}过程信息','{}结果信息','{}日志信息']
        data_group_feature_desc = ['{}，输入时间','{}，校验时间','{}，处理时间','{}，返回码，返回时间','{}，日志时间，线程ID']
        #进度条功能
        i = 0
        widgets = ['Progress: ', FormatLabel('%(value)d of %(max)d '), Percentage(), ' ', Bar('#'),' ', Timer(), ' ', ETA(), ' ', FileTransferSpeed()]
        pbar = ProgressBar(widgets=widgets, maxval=len(self.catalog_names)).start()     #进度条功能
        pbar.update_interval = 1
        
        for catalog_name, catalog_value in self.catalog_names.items():
            sect_catalog_name = '{}.{}'.format(catalog_name, catalog_value)
            if len(self.catalog_sections[sect_catalog_name]) == 0:
                continue
            function_name = catalog_value.strip()
            for elem in self.catalog_sections[sect_catalog_name]['业务要素']:
                elem = elem.strip().split('：')
                if elem[0].strip() == '输入要素':
                    in_elem = elem[1].strip()
                    data_group = in_elem.split(',')[0].strip()
                elif elem[0].strip() == '输出要素':
                    out_elem = elem[1].strip()
                else:
                    print('要素line: %s format error.' %('：'.join(elem)))
            #function_desc = self.catalog_sections[sect_catalog_name]['功能要求'][0].strip().split(',')[-1]
            start_cell = 'E{}'.format(write_row_cnt + 1)
            for loop_id in range(0, 5):
                write_row_cnt = write_row_cnt + 1
                self.ws.insert_rows(write_row_cnt)
                self.duplicateRowStyle(self.ws, self.ws[2], write_row_cnt-2)
                self.ws['E{}'.format(write_row_cnt)] = function_name
                self.ws['F{}'.format(write_row_cnt)] = func_desc[loop_id].format(function_name)
                self.ws['G{}'.format(write_row_cnt)] = data_type[loop_id]
                self.ws['H{}'.format(write_row_cnt)] = data_group_desc[loop_id].format(function_name)
                self.ws['I{}'.format(write_row_cnt)] = data_group_feature_desc[loop_id].format(in_elem)
                self.ws['J{}'.format(write_row_cnt)] = 1
                self.ws['K{}'.format(write_row_cnt)] = ""
            end_cell = 'E{}'.format(write_row_cnt)
            self.ws.merge_cells('{}:{}'.format(start_cell, end_cell))
            
            #更新进度
            i = i + 1
            pbar.update(i)
        self.wb.save("cosmic_tt3.xlsx")
            
    def showResult(self):
        for catalog_name, catalog_vale in self.catalog_names.items():
            print("%s->%s" %(catalog_name, catalog_vale))
        for catalog_name, section_list in self.catalog_sections.items():
            for section_name, section_values in section_list.items():
                for section_value in section_values:
                    print("%s_%s = %s" %(catalog_name, section_name, section_value))

def Usage(command):
    print("usage:" + command + " excel_file [txt_file]")
    print("example: " + command + " *.cosmic.*")
    print("example: " + command + " 12345.cosmic.1 12345.cosmic.2")

if __name__ == '__main__':
    if len(sys.argv) == 1:
        Usage(sys.argv[0])
    else:
        try:
            client = CCosmic()
            client.processAllFiles()
            #client.showResult()
            time.sleep(0)
        except KeyboardInterrupt as e:
            print(e)
        except IOError as e:
            print(e)
        except ValueError as e:
            print(e)
        finally:
            pass