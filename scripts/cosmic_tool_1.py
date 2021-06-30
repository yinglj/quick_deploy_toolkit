#!/usr/bin/env python
# -*- coding: utf-8 -*-

'''
name:
cosmic_tool_1.py
email:
yinglj@asiainfo.com
title:
cosmic_tool_1.py for Linux
'''

import os
import time
import configparser
import re
import string
import logging
import sys
import cmd
import subprocess
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment
from progressbar import *

class CCosmic:
    def __init__(self):
        self.listPstackFiles = []
        self.listStat = []
        self.totalThreads = 0
        self.listIgnores = {"billing billing" : 0, 
                            "exe =":0, 
                            "host:":0, 
                            "Connection":0, 
                            "static_core.txt":0, 
                            "total threads":0}
        self.listShowIgnores = {"exe =":[]} #用于显示明细
        self.backtrace = ""
        self.mapStatic = {}
        self.mapBusiStatic = {}

    def processAllFiles(self):
        for i in range(1, len(sys.argv)):
            self.listPstackFiles.append(sys.argv[i])
            # print "listPstackFiles", self.listPstackFiles;
        for v in self.listPstackFiles:
            self.processOneFile(v)
        self.showResult()

    #复制单元格格式
    def duplicateCellStyle(self, cell1, cell2):
        from copy import copy
        cell1.font = copy(cell2.font)
        cell1.border = copy(cell2.border)
        cell1.fill = copy(cell2.fill)
        cell1.number_format = copy(cell2.number_format)
        cell1.protection = copy(cell2.protection)
        cell1.alignment = copy(cell2.alignment)

    #复制整行单元格格式，从row行复制到row+1行
    def duplicateRowStyle(self, ws, row, seqno=1):
        for i in range(ws.min_column-1, ws.max_column-1):
            cell1 = '{}{}'.format(chr(ord('A')+i), row[0].row+seqno)
            self.duplicateCellStyle(ws[cell1], row[i])   #复制单元格格式
            
    def processOneFile(self, excel_file):
        if not os.path.exists(excel_file):
            return True
        # 加载文件
        wb = load_workbook(excel_file)
        # 读取sheetname
        print('输出文件所有工作表名：\n', wb.sheetnames)
        SHEET_NAME = 'Sheet1'
        START_ROW = 2 #从第二行开始解析
        MERGE_COL = 5   #合并单元格到第5列
        ws = wb[SHEET_NAME]
        ws_copy = wb.copy_worksheet(ws) #copy 整个sheet
        # 按照行列操作
        block_rows = 0
        insertCount = 0
        row1=row2=row3=ws[1]

        #进度条功能
        i = 0
        widgets = ['Progress: ', FormatLabel('%(value)d of %(max)d '), Percentage(), ' ', Bar('#'),' ', Timer(), ' ', ETA(), ' ', FileTransferSpeed()]
        pbar = ProgressBar(widgets=widgets, maxval=ws.max_row).start()     #进度条功能
        pbar.update_interval = 1
    
        for row in ws.iter_rows(min_row=START_ROW, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if row[4].value is not None:
                if block_rows == 3:
                    #在原sheet 数据块的第三行后插入
                    INSERT_SEQNO1 = 1   #表示第1行后
                    ws_copy.insert_rows(row1[0].row+insertCount+INSERT_SEQNO1)
                    self.duplicateRowStyle(ws_copy, row1, insertCount+INSERT_SEQNO1)
                    ws_copy['F{}'.format(row1[0].row+insertCount+INSERT_SEQNO1)] = "读取并解析{}".format(row1[5].value)
                    ws_copy['G{}'.format(row1[0].row+insertCount+INSERT_SEQNO1)] = "R"
                    ws_copy['H{}'.format(row1[0].row+insertCount+INSERT_SEQNO1)] = row1[7].value
                    ws_copy['I{}'.format(row1[0].row+insertCount+INSERT_SEQNO1)] = row1[8].value
                    ws_copy['J{}'.format(row1[0].row+insertCount+INSERT_SEQNO1)] = 1
                    ws_copy['K{}'.format(row1[0].row+insertCount+INSERT_SEQNO1)] = ""
                    insertCount = insertCount+1

                    #在原sheet 数据块的第三行后插入
                    INSERT_SEQNO3 = 3   #表示第3行后
                    ws_copy.insert_rows(row1[0].row+insertCount+INSERT_SEQNO3)
                    self.duplicateRowStyle(ws_copy, row1, insertCount+INSERT_SEQNO3)
                    ws_copy['F{}'.format(row1[0].row+insertCount+INSERT_SEQNO3)] = "将{}记录到日志文件中".format(row1[4].value)
                    ws_copy['G{}'.format(row1[0].row+insertCount+INSERT_SEQNO3)] = "W"
                    ws_copy['H{}'.format(row1[0].row+insertCount+INSERT_SEQNO3)] = row1[7].value
                    ws_copy['I{}'.format(row1[0].row+insertCount+INSERT_SEQNO3)] = row3[8].value+",成功/失败标识,系统时间"
                    ws_copy['J{}'.format(row1[0].row+insertCount+INSERT_SEQNO3)] = 1
                    ws_copy['K{}'.format(row1[0].row+insertCount+INSERT_SEQNO3)] = ""
                    insertCount = insertCount+1
                row1 = row
                block_rows=1
            else:
                block_rows = block_rows + 1
                if block_rows == 2: #第二行
                    row2 = row
                elif block_rows == 3:   #第三行
                    row3 = row
            #更新进度
            i = i + 1
            pbar.update(i)
        
        #处理最后一个数据块
        if block_rows == 3:
            #在原sheet 数据块的第三行后插入
            INSERT_SEQNO1 = 1   #表示第1行后
            ws_copy.insert_rows(row1[0].row+insertCount+INSERT_SEQNO1)
            self.duplicateRowStyle(ws_copy, row1, insertCount+INSERT_SEQNO1)
            ws_copy['F{}'.format(row1[0].row+insertCount+INSERT_SEQNO1)] = "读取并解析{}".format(row1[5].value)
            ws_copy['G{}'.format(row1[0].row+insertCount+INSERT_SEQNO1)] = "R"
            ws_copy['H{}'.format(row1[0].row+insertCount+INSERT_SEQNO1)] = row1[7].value
            ws_copy['I{}'.format(row1[0].row+insertCount+INSERT_SEQNO1)] = row1[8].value
            ws_copy['J{}'.format(row1[0].row+insertCount+INSERT_SEQNO1)] = 1
            ws_copy['K{}'.format(row1[0].row+insertCount+INSERT_SEQNO1)] = ""
            insertCount = insertCount+1

            #在原sheet 数据块的第三行后插入
            INSERT_SEQNO3 = 3   #表示第3行后
            ws_copy.insert_rows(row1[0].row+insertCount+INSERT_SEQNO3)
            self.duplicateRowStyle(ws_copy, row1, insertCount+INSERT_SEQNO3)
            ws_copy['F{}'.format(row1[0].row+insertCount+INSERT_SEQNO3)] = "将{}记录到日志文件中".format(row1[4].value)
            ws_copy['G{}'.format(row1[0].row+insertCount+INSERT_SEQNO3)] = "W"
            ws_copy['H{}'.format(row1[0].row+insertCount+INSERT_SEQNO3)] = row1[7].value
            ws_copy['I{}'.format(row1[0].row+insertCount+INSERT_SEQNO3)] = row3[8].value+",成功/失败标识,系统时间"
            ws_copy['J{}'.format(row1[0].row+insertCount+INSERT_SEQNO3)] = 1
            ws_copy['K{}'.format(row1[0].row+insertCount+INSERT_SEQNO3)] = ""
            insertCount = insertCount+1
        
        #ws_copy按列合并单元格
        for col in ws_copy.iter_cols(min_row=START_ROW, max_row=ws_copy.max_row, min_col=1, max_col=MERGE_COL):
            start_cell=end_cell=""
            for cell in col:
                if cell.value is not None:
                    if end_cell != start_cell:
                        ws_copy.merge_cells('{}:{}'.format(start_cell, end_cell))
                    start_cell = cell.coordinate
                else:
                    end_cell = cell.coordinate
            if end_cell != start_cell:
                ws_copy.merge_cells('{}:{}'.format(start_cell, end_cell))

        wb.save('./scripts/mdb2.xlsx')  # 保存更改
        return True

    def showResult(self):
        return True

def Usage(command):
    print("usage:" + command + " [file]")
    print("example: " + command + " *.cosmic.*")
    print("example: " + command + " 12345.cosmic.1 12345.cosmic.2")

'''
基础Table类分析功能	传入调用分析结构参数	E	分析标识结构	分析标识ID	1 
	Table类分析功能逻辑进行封装处理	R	分析对象	分析对象，表名，字段类型，字段名	1 
	Table类调用分析结果封装并进行校验	X	分析返回结果	分析返回值	1 
基础Table类后备功能	传入调用后备结构参数	E	后备标识结构	后备标识ID	1 
	Table类后备功能逻辑进行封装处理	R	后备对象	后备对象，表名，字段类型，字段名	1 
	Table类调用后备结果封装并进行校验	X	后备返回结果	后备返回值	1 
基础Table类备份解锁功能	传入调用备份解锁结构参数	E	备份解锁标识结构	备份解锁标识ID	1 
	Table类备份解锁功能逻辑进行封装处理	R	备份解锁对象	备份解锁对象，表名，字段类型，字段名	1 
	Table类调用备份解锁结果封装并进行校验	X	备份解锁返回结果	备份解锁返回值	1
'''
#!扩展5行模式
if __name__ == '__main__':
    if len(sys.argv) == 1:
        Usage(sys.argv[0])
    else:
        try:
            client = CCosmic()
            client.processAllFiles()
            time.sleep(0)
        except KeyboardInterrupt as e:
            print(e)
        except IOError as e:
            print(e)
        except ValueError as e:
            print(e)
        finally:
            pass